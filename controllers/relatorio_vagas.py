from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import date
import re


class Planilha:
    def __init__(self) -> None:
        """Instancia a classe com as configurações iniciais,
        para criação e ativação do arquivo excel.
        """
        self.arquivo_excel = Workbook()
        self.sheetVagas = self.arquivo_excel.active

    def formatar_planilha(self) -> None:
        """Função responsável pela formatação da planilha,
            ela coloca a guia com a data atual e renomeia as colunas para 
            receber os novos parâmetros, e coloca quebra de linha.
        """
        self.sheetVagas.title = format(date.today())

        self.sheetVagas['A1'] = "Nome"
        self.sheetVagas['B1'] = "Local"
        self.sheetVagas['C1'] = "Descrição"

        self.sheetVagas['A1'].font = Font(
            size=18, bold=True)
        self.sheetVagas['B1'].font = Font(
            size=18, bold=True)
        self.sheetVagas['C1'].font = Font(
            size=18, bold=True)

        self.sheetVagas['A1'].alignment = Alignment(
            horizontal='center', vertical='center')
        self.sheetVagas['B1'].alignment = Alignment(
            horizontal='center', vertical='center')
        self.sheetVagas['C1'].alignment = Alignment(
            horizontal='center', vertical='center')

        self.sheetVagas.column_dimensions['A'].width = 50
        self.sheetVagas.column_dimensions['B'].width = 20
        self.sheetVagas.column_dimensions['C'].width = 70
        self.arquivo_excel.save('vagas.xlsx')

    def escreve_dados(self, dados_vagas: dict) -> None:
        """Função resonsável por inserir os dados na planilha.
        Args:
            dados_vagas (dict): recebe um dicionário contendo as vagas
        """
        i = 0
        for vaga in dados_vagas:
            self.sheetVagas.cell(row=i+2, column=1).value = vaga
            self.sheetVagas.cell(
                row=i+2, column=2).value = dados_vagas[vaga]['local']
            descricao = re.sub('\n', ' ', dados_vagas[vaga]['descricao'])
            self.sheetVagas.cell(row=i+2, column=3).value = descricao
            self.sheetVagas[f'A{i+2}'].alignment = Alignment(vertical='center')
            self.sheetVagas[f'B{i+2}'].alignment = Alignment(vertical='center')
            self.sheetVagas[f'C{i+2}'].alignment = Alignment(wrap_text=True)
            i += 1
        self.arquivo_excel.save('vagas.xlsx')
